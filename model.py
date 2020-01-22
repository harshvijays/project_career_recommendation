# -*- coding: utf-8 -*-
"""
Created on Tue Jan 21 16:17:26 2020

@author: DBDA
"""

import pandas as pd
import pickle
#df = pd.read_excel("newprojectdataset.xlsx")

df = pd.read_excel("Book1.xlsx")
dum_df = pd.get_dummies(df)
#dum_df = dum_df.drop('Class_Benign', axis=1)

from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report, accuracy_score
#from sklearn.ensemble import BaggingClassifier

X = dum_df.iloc[:,0:20]
y = dum_df.iloc[:,20]

X_train, X_test, y_train, y_test = train_test_split(X, y,test_size = 0.3,
                                                    random_state=2018,
                                                    stratify=y)
#
from sklearn.ensemble import GradientBoostingClassifier
#
## Create training and test sets
#X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.3,
#                                                    random_state=2018)
#
clf = GradientBoostingClassifier(random_state=1200)
import openpyxl

wb_obj = openpyxl.load_workbook("Test1.xlsx")
sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row = 1, column = 1)
print(cell_obj.value)
max_row=sheet_obj.max_row
print(max_row)
max_col=sheet_obj.max_column
clf.fit(X_train,y_train)

pickle.dump(clf,open('model.pkl','wb'))

model=pickle.load(open('model.pkl','rb'))


import numpy as np

dtest=pd.read_excel("Test1.xlsx")

#y_pred = clf.predict(X_test)
y_pred = clf.predict(dtest)
print(y_pred)
print("\n---------using pickle prediction : ",model.predict(dtest))
y_pred = clf.predict(X_test)
print(confusion_matrix(y_test, y_pred))
print(classification_report(y_test, y_pred))
print(accuracy_score(y_test,y_pred))
                                         
